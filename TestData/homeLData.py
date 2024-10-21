import openpyxl
class HomeLData:

    #test_homeL_page_data = [{"Name":"Taranjit","Email":"taranjit@tata.com", "password":"pass@3212"}]

    @staticmethod
    def getTestData(test_case_name, sheet_name):
        Dict = {}
        # Load the workbook
        book = openpyxl.load_workbook(r"D:\PytestPython\DummySeleniumProject\TestData\Book1.xlsx")

        # Access the specified sheet by name
        sheet = book[sheet_name]

        # Loop through all rows to find the matching test case
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                # Loop through all columns of that row to gather the test data
                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        # Return the dictionary with test data
        return [Dict]