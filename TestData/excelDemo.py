import openpyxl
import sys

def getData(sheet_name,test_name):
    book = openpyxl.load_workbook(r"D:\PytestPython\DummySeleniumProject\TestData\Book1.xlsx")


    sheet = book[sheet_name]

    cell = sheet.cell(row=1,column=2)

    # print(cell.value)
    #
    # print(f"{sheet.cell(row=2,column=2).value}")
    #
    # print(sheet.max_row)
    #
    # print(sheet.max_column)
    #
    # print(sheet['A3'].value)
    dict = {}

    for i in range(1,sheet.max_row+1):
        if sheet.cell(row=i,column=1).value==test_name:
            for j in range(2,sheet.max_column+1):
               #dict['firstname]='taran example'
               dict[sheet.cell(row=1,column=j).value]= sheet.cell(row=i,column=j).value

    print(dict)
sys.path.append("D:/PytestPython/E2EPythonSelenium")
sys.path.append("D:/PytestPython/E2EPythonSelenium")
getData("Sheet2","testCase1")
getData("Sheet2","testCase2")
getData("Sheet2","testCase3")
