import time

import pytest
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from trio import WouldBlock

from utilities.BaseClass import BaseClass


class TestUploadDownload(BaseClass):

    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get('https://rahulshettyacademy.com/upload-download-test/index.html')
    file_path="D:\\PytestPython\\download.xlsx"
    wait = WebDriverWait(driver,25)
    def test_upload_download(self):
        try:
            fruit_name = 'Apple'
            Dict={}
            file = openpyxl.load_workbook('D:\\PytestPython\\download.xlsx')
            sheet=file.active
            sheet.cell(row=Dict["row"],column=Dict["col"]).value='700'

            for i in range(1,sheet.max_column+1):
                if sheet.cell(row=1,column=i).value == 'price':
                        Dict["col"]=i

            for i in range(1,sheet.max_row+1):
                for j in range(1,sheet.max_column+1):
                    if sheet.cell(row=i,column=j).value=="Apple":
                        Dict["row"]=i

            sheet.cell(row=Dict["row"], column=Dict["col"]).value = '700'

            log= self.getLogger()
            #TestUploadDownload.driver.implicitly_wait(15)
            downloadButton = (By.ID,"downloadButton")
            #download the file
            TestUploadDownload.wait.until(EC.visibility_of_element_located(downloadButton))
            TestUploadDownload.driver.find_element(*downloadButton).click()
            # TestUploadDownload.driver.refresh()
            #upload the file
            files_input=TestUploadDownload.driver.find_element(By.CSS_SELECTOR,"input[type='file']")
            files_input.send_keys(TestUploadDownload.file_path)

            #fileUploaded message
            #toast_locator= (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
            toast_locator = (By.XPATH,"//div[contains(text(),'Updated Excel Data Successfully.')]")
            TestUploadDownload.wait.until(EC.visibility_of_element_located(toast_locator))
            message = TestUploadDownload.driver.find_element(*toast_locator).text

            #message = TestUploadDownload.driver.execute_script("return arguments[0].innerText;", toast_locator)
            log.info(f"File uploaded successfully: {message}")
            print(f'{message}')
            time.sleep(10)
            price_attribute = TestUploadDownload.driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
            actualPriceText=TestUploadDownload.driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_attribute+"-undefined']").text
            log.info(f"The actual price of the {fruit_name} = {actualPriceText}")
            TestUploadDownload.driver.close()
        except Exception as e:
            log.exception(f"{e}")